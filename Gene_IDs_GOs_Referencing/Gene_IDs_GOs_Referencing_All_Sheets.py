# Import modules
import pandas as pd
import openpyxl

# Disable warning when replacing "GO:" with an empty string
pd.options.mode.chained_assignment = None  # default='warn'.

# Convert the excel file into a dictionary where the keys are the sheet names and the values are the resulting dataframes
Sheet_dict = pd.read_excel("EggNog_v360_proteome_20230615_HopBA1_RNAseq_DEseq20230614.xlsx", 
                        sheet_name = None)

# Iterate through the dictionary and write a new excel file containing only the gene IDs and associated GO terms for each sheet from the original spreadsheet
for sheet_name, df in Sheet_dict.items():

    if sheet_name == "Eggnog NbLAB360":
        df = df[df.GOs != "-"]
        df['GOs'] = df['GOs'].str.replace("GO:", "")
        with pd.ExcelWriter('Extracted_EggNog_NbLAB360_IDs_GOs.xlsx', mode="a") as writer:
            df.to_excel(writer, columns=["query", "GOs"], sheet_name="Eggnog NbLAB360")

    elif sheet_name == "Readme":
        continue

    elif sheet_name == "Leaves UP with HopBA1":
        df = df[df.GOs != "-"]
        df['GOs'] = df['GOs'].str.replace("GO:", "")
        with pd.ExcelWriter('Extracted_EggNog_NbLAB360_IDs_GOs.xlsx', mode="a") as writer:
            df.to_excel(writer, columns=["query", "GOs"], sheet_name="Leaves UP with HopBA1")

    elif sheet_name == "Leaves DOWN":
        df = df[df.GOs != "-"]
        df['GOs'] = df['GOs'].str.replace("GO:", "")
        with pd.ExcelWriter('Extracted_EggNog_NbLAB360_IDs_GOs.xlsx', mode="a") as writer: # mode="a" to append to an existing spreadsheet
            df.to_excel(writer, columns=["query", "GOs"], sheet_name="Leaves DOWN")

    elif sheet_name == "Petioles UP":
        df = df[df.GOs != "-"]
        df['GOs'] = df['GOs'].str.replace("GO:", "")
        with pd.ExcelWriter('Extracted_EggNog_NbLAB360_IDs_GOs.xlsx', mode="a") as writer:
            df.to_excel(writer, columns=["query", "GOs"], sheet_name="Petioles UP")

    elif sheet_name == "Petioles DOWN":
        df = df[df.GOs != "-"]
        df['GOs'] = df['GOs'].str.replace("GO:", "")
        with pd.ExcelWriter('Extracted_EggNog_NbLAB360_IDs_GOs.xlsx', mode="a") as writer:
            df.to_excel(writer, columns=["query", "GOs"], sheet_name="Petioles DOWN")

# Load the previously generated workbook, get its sheet names and start a sheet counter
wb = openpyxl.load_workbook("Extracted_EggNog_NbLAB360_IDs_GOs.xlsx")
sheet_names = wb.sheetnames
sheet_count = None

# Iterate through the worksheets
for sheet in wb.worksheets:

    # Count the number of sheets iterated through
    if sheet_count is None:
        sheet_count = 0
    else:
        sheet_count += 1

    # Iterate through the rows returning row objects
    rows = sheet.iter_rows(min_row = 2, min_col = 2) # Desired data starts from row and column 2

    # Append the IDs to a dictionary as keys and the GO terms as values after splitting the latter by commas.
    ID_GO_Dict = {}
    for ID,GOs in rows:
        ID_GO_Dict[ID.value] = str(GOs.value).split(",")

    # Create a BinGO annotation file for UP genes in leaves specifying the species name, type of GO terms and the BinGO curator in the first line
    with open("Nb_GO_Terms_EggNog_NbLAB360_" + str(sheet_names[sheet_count]) + "_HopBA1_RNAseq.txt", "w") as f:
        print("(species=Nicotiana benthamiana)(type=Biological Process)(curator=GO)", file=f)    

    # Iterate through the dictionary to write to the BinGO annotation file Nb's gene IDs and their associated GO terms
    for ID,GOs in ID_GO_Dict.items():
        for GO in GOs:
            if ID and GO is not None:
                with open("Nb_GO_Terms_EggNog_NbLAB360_" + str(sheet_names[sheet_count]) + "_HopBA1_RNAseq.txt","a") as f:
                    print(ID + " = " + str(GO), file=f)

print("####Done!####")
