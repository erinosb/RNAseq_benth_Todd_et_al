# Import modules
import pandas as pd
import openpyxl

# Disable warning when replacing "ko:" with an empty string
pd.options.mode.chained_assignment = None  # default='warn'.

# Convert the excel file into a dictionary where the keys are the sheet names and the values are the resulting dataframes
Sheet_dict = pd.read_excel("EggNog_v360_proteome_20230615_HopBA1_RNAseq_DEseq20230614.xlsx", 
                        sheet_name = None)

# Iterate through the dictionary and write a new excel file containing only the gene IDs and associated KEGG IDs for each sheet from the original spreadsheet
for sheet_name, df in Sheet_dict.items():

    if sheet_name == "Eggnog NbLAB360_Original":
        continue

    elif sheet_name == "Eggnog NbLAB360":
        df = df[df.KEGG_ko != "-"]
        df["KEGG_ko"] = df["KEGG_ko"].str.replace("ko:", "")
        with pd.ExcelWriter('Extracted_EggNog_NbLAB360_KEGG_IDs.xlsx') as writer: # mode="a" to append to an existing spreadsheet
            df.to_excel(writer, columns=["query", "KEGG_ko"], sheet_name="Eggnog NbLAB360")

    elif sheet_name == "Readme":
        continue

    elif sheet_name == "Leaves UP with HopBA1":
        df = df[df.KEGG_ko != "-"]
        df["KEGG_ko"] = df["KEGG_ko"].str.replace("ko:", "")
        with pd.ExcelWriter('Extracted_EggNog_NbLAB360_KEGG_IDs.xlsx', engine='openpyxl',mode="a") as writer: # mode="a" to append to an existing spreadsheet
            df.to_excel(writer, columns=["query", "KEGG_ko"], sheet_name="Leaves UP with HopBA1")

    elif sheet_name == "Leaves DOWN":
        df = df[df.KEGG_ko != "-"]
        df["KEGG_ko"] = df["KEGG_ko"].str.replace("ko:", "")
        with pd.ExcelWriter('Extracted_EggNog_NbLAB360_KEGG_IDs.xlsx', engine='openpyxl', mode="a") as writer:
            df.to_excel(writer, columns=["query", "KEGG_ko"], sheet_name="Leaves DOWN")

    elif sheet_name == "Petioles UP":
        df = df[df.KEGG_ko != "-"]
        df["KEGG_ko"] = df["KEGG_ko"].str.replace("ko:", "")
        with pd.ExcelWriter('Extracted_EggNog_NbLAB360_KEGG_IDs.xlsx', engine='openpyxl', mode="a") as writer:
            df.to_excel(writer, columns=["query", "KEGG_ko"], sheet_name="Petioles UP")

    elif sheet_name == "Petioles DOWN":
        df = df[df.KEGG_ko != "-"]
        df["KEGG_ko"] = df["KEGG_ko"].str.replace("ko:", "")
        with pd.ExcelWriter('Extracted_EggNog_NbLAB360_KEGG_IDs.xlsx', engine='openpyxl', mode="a") as writer:
            df.to_excel(writer, columns=["query", "KEGG_ko"], sheet_name="Petioles DOWN")

# Load the previously generated workbook, get its sheet names and start a sheet counter
wb = openpyxl.load_workbook("Extracted_EggNog_NbLAB360_KEGG_IDs.xlsx")
sheet_names = wb.sheetnames
sheet_count = None

# Iterate through the worksheets
for sheet in wb.worksheets:

    # Count the number of sheets iterated through
    if sheet_count is None:
        sheet_count = 0
    else:
        sheet_count += 1

    # Iterate through the rows returning row objects
    rows = sheet.iter_rows(min_row = 2, min_col = 2) # Desired data starts from row and column 2

    # Append the gene IDs to a dictionary as keys and the KEGG IDs as values after splitting the latter by commas.
    ID_KEGG_Dict = {}
    for ID,KEGGs in rows:
        ID_KEGG_Dict[ID.value] = str(KEGGs.value).split(",")

    # Iterate through the dictionary to write to the KEGG annotation file
    for ID,KEGGs in ID_KEGG_Dict.items():
        for KEGG in KEGGs:
            if ID and KEGG is not None:
                if sheet_count == 0:
                    
                    # Generate a .txt file containing the association of gene IDs and KEGG IDs for the whole proteome
                    with open("Nb_KEGG_IDs_EggNog_NbLAB360_HopBA1_RNAseq.txt","a") as f:
                        print(ID + " = " + str(KEGG), file=f)
                    
                    # Generate a .txt file containing only the KEGG IDs for the whole proteome
                    with open("Nb_KEGG_IDs_Only_EggNog_NbLAB360_HopBA1_RNAseq.txt", "a") as f:
                        print(KEGG, file=f)

                else:
                    # Generate a .txt file containing only the KEGG IDs for each remaining sheets. Therefore, these are the KEGG IDs that have at least one DEG associated.
                    with open("Nb_KEGG_IDs_EggNog_NbLAB360_" + str(sheet_names[sheet_count]) + "_Only_HopBA1_RNAseq.txt","a") as f:
                        print(KEGG, file=f)

print("####Done!####")
